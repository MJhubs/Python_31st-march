# In[6]:


import openpyxl as op

wb = op.load_workbook(r"C:\Users\Meet\Downloads\realestatedata.xlsx")
print(type(wb))


# In[7]:


sheets = wb.sheetnames
print("List of sheets")


# In[8]:


for x in sheets:
    print(x)
    sheet_name = x


# In[9]:


sheet = wb[sheet_name]
print(type(sheet))


# In[10]:


print(sheet.title)


# In[11]:


sheets = wb.sheetnames
print("listing of sheets ")


# In[12]:


for i in sheets:
    print(i)
    sheet = wb[sheet_name]

print(type(sheet))

print(sheet.title)


# In[13]:


#version2
data1 = sheet['C4']

print(data1)
print(data1.value)

for i in range(1, 190):
    #print(sheet.cell(row=i, column=10).value)
    price = sheet.cell(row=i, column=4).value;
    #print(price);
    houseType = sheet.cell(row=i, column=1).value;
    #print(houseType);
    description = sheet.cell(row=i, column=2).value;
    #print(description);
    numberBedrooms = sheet.cell(row=i, column=3).value;
    #print(numberBedrooms);
    houseprice = (houseType, description, numberBedrooms, price)
    print(houseprice)


# In[14]:


#version3
sheets = wb.sheetnames
print("listing of sheets ")
for i in sheets:
    print(i)

sheet = wb[sheet_name]

print(type(sheet))

print(sheet.title)

data1 = sheet['C4']

print(data1)
print(data1.value)

houseprice_1 = []
for i in range(1, 190):
    #print(sheet.cell(row=i, column=10).value)
    price = sheet.cell(row=i, column=4).value;
    #print(price);
    houseType = sheet.cell(row=i, column=1).value;
    #print(houseType);
    description = sheet.cell(row=i, column=2).value;
    #print(description);
    numberBedrooms = sheet.cell(row=i, column=3).value;
    #print(numberBedrooms);
    houseprice = (houseType, description, numberBedrooms, price)
    print(houseprice)
    houseprice_1.append(houseprice)

for i in houseprice_1:
    print(i)


# In[15]:


#version 4
sheets = wb.sheetnames
print("listing of sheets ")
for i in sheets:
    print(i)

sheet = wb[sheet_name]

print(type(sheet))

print(sheet.title)

data1 = sheet['C4']

print(data1)
print(data1.value)


houseprice_1 = []
for i in range(1, 190):
    #print(sheet.cell(row=i, column=10).value)
    price = sheet.cell(row=i, column=4).value;
    #print(price);
    houseType = sheet.cell(row=i, column=1).value;
    #print(houseType);
    description = sheet.cell(row=i, column=2).value;
    #print(description);
    numberBedrooms = sheet.cell(row=i, column=3).value;
    #print(numberBedrooms);
    houseprice = (houseType, description, numberBedrooms, price)
    print(houseprice)
    houseprice_1.append(list(houseprice))

for i in houseprice_1:
    print(i)
# printing only house type is detached
for i in houseprice_1:
    if i[0] == 'Detached':
        print(i)
